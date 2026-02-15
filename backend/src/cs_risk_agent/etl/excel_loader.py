"""Excel ローダー - 財務データExcelファイルの読み込み・バリデーション.

ユーザーがアップロードしたExcelファイルから財務データを読み込み、
スキーマバリデーションを実施して構造化データに変換する。
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import structlog

logger = structlog.get_logger(__name__)

# 期待される列名マッピング（日本語 -> 英語キー）
COLUMN_MAP: dict[str, str] = {
    # 基本情報
    "企業名": "company_name",
    "企業コード": "company_code",
    "EDINET コード": "edinet_code",
    "EDINETコード": "edinet_code",
    "証券コード": "securities_code",
    "業種": "industry_name",
    "業種コード": "industry_code",
    "会計年度": "fiscal_year",
    "四半期": "fiscal_quarter",
    # 売上・収益
    "売上高": "revenue",
    "売上原価": "cost_of_sales",
    "営業収益": "operating_revenue",
    # 利益
    "売上総利益": "gross_profit",
    "営業利益": "operating_income",
    "経常利益": "ordinary_income",
    "当期純利益": "net_income",
    "親会社株主に帰属する当期純利益": "net_income_parent",
    # 資産
    "総資産": "total_assets",
    "流動資産": "current_assets",
    "固定資産": "noncurrent_assets",
    "現金及び預金": "cash_and_deposits",
    "売掛金": "accounts_receivable",
    "受取手形及び売掛金": "accounts_receivable",
    "棚卸資産": "inventories",
    # 負債
    "負債合計": "total_liabilities",
    "流動負債": "current_liabilities",
    "固定負債": "noncurrent_liabilities",
    "短期借入金": "short_term_loans",
    "長期借入金": "long_term_loans",
    # 純資産
    "純資産": "net_assets",
    "株主資本": "shareholders_equity",
    "自己資本比率": "equity_ratio",
    # キャッシュフロー
    "営業CF": "cf_operating",
    "営業活動によるキャッシュ・フロー": "cf_operating",
    "投資CF": "cf_investing",
    "投資活動によるキャッシュ・フロー": "cf_investing",
    "財務CF": "cf_financing",
    "財務活動によるキャッシュ・フロー": "cf_financing",
    # その他
    "従業員数": "employees",
    "配当金": "dividends",
    "EPS": "eps",
    "BPS": "bps",
    "PER": "per",
    "PBR": "pbr",
    "ROE": "roe",
    "ROA": "roa",
}

# 必須列リスト
REQUIRED_COLUMNS: set[str] = {
    "company_name",
    "fiscal_year",
}


@dataclass
class ValidationError:
    """バリデーションエラー."""

    row: int | None
    column: str | None
    message: str
    severity: str = "error"  # error, warning


@dataclass
class LoadResult:
    """Excel読み込み結果."""

    records: list[dict[str, Any]] = field(default_factory=list)
    total_rows: int = 0
    valid_rows: int = 0
    errors: list[ValidationError] = field(default_factory=list)
    warnings: list[ValidationError] = field(default_factory=list)
    column_mapping: dict[str, str] = field(default_factory=dict)
    metadata: dict[str, Any] = field(default_factory=dict)

    @property
    def is_valid(self) -> bool:
        """エラーなしの場合True."""
        return len(self.errors) == 0


class ExcelLoader:
    """Excel ローダー.

    openpyxl を使用してExcelファイルを読み込み、
    列名マッピングとバリデーションを実行する。
    """

    def __init__(
        self,
        column_map: dict[str, str] | None = None,
        required_columns: set[str] | None = None,
        sheet_name: str | int = 0,
    ) -> None:
        """初期化.

        Args:
            column_map: カスタム列名マッピング。
            required_columns: 必須列名セット。
            sheet_name: 読み込みシート名またはインデックス。
        """
        self._column_map = column_map or COLUMN_MAP
        self._required_columns = required_columns or REQUIRED_COLUMNS
        self._sheet_name = sheet_name

    def load(self, file_path: Path | str) -> LoadResult:
        """Excelファイルを読み込み.

        Args:
            file_path: Excelファイルパス。

        Returns:
            読み込み結果。
        """
        file_path = Path(file_path)
        result = LoadResult()

        logger.info("excel_loader.start", file_path=str(file_path))

        if not file_path.exists():
            result.errors.append(
                ValidationError(
                    row=None,
                    column=None,
                    message=f"ファイルが見つかりません: {file_path}",
                )
            )
            return result

        if file_path.suffix not in (".xlsx", ".xlsm", ".xls"):
            result.errors.append(
                ValidationError(
                    row=None,
                    column=None,
                    message=f"サポートされていないファイル形式: {file_path.suffix}",
                )
            )
            return result

        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, read_only=True, data_only=True)

            # シート取得
            if isinstance(self._sheet_name, int):
                if self._sheet_name >= len(wb.sheetnames):
                    result.errors.append(
                        ValidationError(
                            row=None,
                            column=None,
                            message=f"シートインデックスが範囲外: {self._sheet_name}",
                        )
                    )
                    wb.close()
                    return result
                ws = wb[wb.sheetnames[self._sheet_name]]
            else:
                if self._sheet_name not in wb.sheetnames:
                    result.errors.append(
                        ValidationError(
                            row=None,
                            column=None,
                            message=f"シートが見つかりません: {self._sheet_name}",
                        )
                    )
                    wb.close()
                    return result
                ws = wb[self._sheet_name]

            # ヘッダー行読み込み
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                result.errors.append(
                    ValidationError(
                        row=None,
                        column=None,
                        message="シートにデータがありません",
                    )
                )
                wb.close()
                return result

            headers = [str(h).strip() if h else "" for h in rows[0]]
            column_mapping = self._map_columns(headers)
            result.column_mapping = column_mapping

            # 必須列チェック
            mapped_columns = set(column_mapping.values())
            missing = self._required_columns - mapped_columns
            if missing:
                result.errors.append(
                    ValidationError(
                        row=None,
                        column=None,
                        message=f"必須列が不足しています: {', '.join(missing)}",
                    )
                )

            # データ行読み込み
            result.total_rows = len(rows) - 1  # ヘッダー除く
            for row_idx, row in enumerate(rows[1:], start=2):
                record = self._parse_row(row, headers, column_mapping, row_idx, result)
                if record:
                    result.records.append(record)
                    result.valid_rows += 1

            wb.close()

            result.metadata = {
                "file": str(file_path),
                "sheet": str(self._sheet_name),
                "columns_detected": len(headers),
                "columns_mapped": len(column_mapping),
            }

        except ImportError:
            result.errors.append(
                ValidationError(
                    row=None,
                    column=None,
                    message="openpyxl がインストールされていません",
                )
            )
        except Exception as e:
            result.errors.append(
                ValidationError(
                    row=None,
                    column=None,
                    message=f"Excel読み込みエラー: {e}",
                )
            )

        logger.info(
            "excel_loader.complete",
            total_rows=result.total_rows,
            valid_rows=result.valid_rows,
            errors=len(result.errors),
            warnings=len(result.warnings),
        )
        return result

    def validate(self, records: list[dict[str, Any]]) -> list[ValidationError]:
        """レコードリストのバリデーション.

        型チェック、範囲チェック、整合性チェックを実行する。

        Args:
            records: バリデーション対象レコードリスト。

        Returns:
            バリデーションエラーリスト。
        """
        errors: list[ValidationError] = []

        for idx, record in enumerate(records, start=1):
            # 会計年度チェック
            fiscal_year = record.get("fiscal_year")
            if fiscal_year is not None:
                if not isinstance(fiscal_year, (int, float)):
                    errors.append(
                        ValidationError(
                            row=idx,
                            column="fiscal_year",
                            message="会計年度は数値である必要があります",
                        )
                    )
                elif fiscal_year < 1900 or fiscal_year > 2100:
                    errors.append(
                        ValidationError(
                            row=idx,
                            column="fiscal_year",
                            message=f"会計年度の値が不正です: {fiscal_year}",
                        )
                    )

            # 財務数値の符号チェック
            positive_fields = [
                "total_assets",
                "revenue",
                "current_assets",
                "employees",
            ]
            for field_name in positive_fields:
                value = record.get(field_name)
                if value is not None and isinstance(value, (int, float)):
                    if value < 0:
                        errors.append(
                            ValidationError(
                                row=idx,
                                column=field_name,
                                message=f"{field_name} が負の値です: {value}",
                                severity="warning",
                            )
                        )

            # 貸借対照表の整合性チェック
            total_assets = record.get("total_assets")
            total_liabilities = record.get("total_liabilities")
            net_assets = record.get("net_assets")
            if all(
                v is not None and isinstance(v, (int, float))
                for v in [total_assets, total_liabilities, net_assets]
            ):
                balance = abs(total_assets - (total_liabilities + net_assets))
                if balance > total_assets * 0.01:  # 1%以上の差異
                    errors.append(
                        ValidationError(
                            row=idx,
                            column="total_assets",
                            message=(
                                f"貸借対照表の不整合: 資産={total_assets}, "
                                f"負債+純資産={total_liabilities + net_assets}"
                            ),
                            severity="warning",
                        )
                    )

        return errors

    def _map_columns(self, headers: list[str]) -> dict[str, str]:
        """ヘッダー名をキー名にマッピング.

        Args:
            headers: Excelヘッダー名リスト。

        Returns:
            ヘッダー名 -> 英語キー名のマッピング。
        """
        mapping: dict[str, str] = {}
        for header in headers:
            if not header:
                continue
            normalized = header.strip()
            if normalized in self._column_map:
                mapping[normalized] = self._column_map[normalized]
            else:
                # 英語キーそのままの場合
                lower = normalized.lower()
                if lower in {v for v in self._column_map.values()}:
                    mapping[normalized] = lower
        return mapping

    def _parse_row(
        self,
        row: tuple[Any, ...],
        headers: list[str],
        column_mapping: dict[str, str],
        row_idx: int,
        result: LoadResult,
    ) -> dict[str, Any] | None:
        """行データを辞書に変換.

        Args:
            row: 行の値タプル。
            headers: ヘッダー名リスト。
            column_mapping: 列名マッピング。
            row_idx: 行番号（エラー報告用）。
            result: 結果オブジェクト（エラー追加用）。

        Returns:
            変換された辞書。全セルが空の場合はNone。
        """
        record: dict[str, Any] = {}
        has_data = False

        for col_idx, value in enumerate(row):
            if col_idx >= len(headers):
                break

            header = headers[col_idx]
            if not header or header not in column_mapping:
                continue

            key = column_mapping[header]
            if value is not None:
                has_data = True
                record[key] = self._coerce_value(value)

        if not has_data:
            return None

        return record

    @staticmethod
    def _coerce_value(value: Any) -> Any:
        """値の型変換.

        Args:
            value: 変換対象の値。

        Returns:
            型変換後の値。
        """
        if value is None:
            return None

        # 文字列の場合: 数値に変換を試行
        if isinstance(value, str):
            cleaned = value.strip().replace(",", "").replace(" ", "")
            if not cleaned:
                return None
            try:
                if "." in cleaned:
                    return float(cleaned)
                return int(cleaned)
            except ValueError:
                return cleaned

        return value
